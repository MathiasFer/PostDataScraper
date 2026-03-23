import pandas as pd
import re
import os
import logging
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright
from posts import InstagramPostFinder
from reels import InstagramReelFinder
from test_existencia import validar_existencia_playwright_batch

# Configuración de logging para seguimiento en consola
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def extract_shortcode(url):
    """Extrae el shortcode de una URL de Instagram."""
    if not isinstance(url, str):
        return None
    # Busca patrones como /p/CODE/, /reel/CODE/, /reels/CODE/
    match = re.search(r'/(?:p|reel|reels)/([^/?#&]+)', url)
    if match:
        return match.group(1)
    return None

def calculate_diff(real, scan):
    """Calcula la diferencia porcentual: abs(real - scan) / max(real, 1) * 100."""
    try:
        real_val = float(real) if real is not None else 0
        scan_val = float(scan) if scan is not None else 0
        if real_val == 0:
            return 0 if scan_val == 0 else 100.0
        return (abs(real_val - scan_val) / max(real_val, 1)) * 100
    except (ValueError, TypeError):
        return 0

def main():
    input_file = 'colima_testeo.xlsx'
    output_final = 'resultados_colima_testeo.xlsx'
    
    if not os.path.exists(input_file):
        logging.error(f"No se encontró el archivo de entrada: {input_file}")
        return

    logging.info(f"Leyendo archivo {input_file}...")
    try:
        df = pd.read_excel(input_file, dtype={'Id Publicación': str})
    except Exception as e:
        logging.error(f"Error al leer el Excel: {e}")
        return
    
    df = df.copy()
    df['index_original'] = list(range(len(df)))
    results = []
    total_rows = len(df)

    if 'Usuario' not in df.columns:
        logging.error("El Excel no tiene la columna 'Usuario'.")
        return

    # Instanciar buscadores (se asume que auth.json existe y es válido)
    post_finder = InstagramPostFinder()
    reel_finder = InstagramReelFinder()

    def infer_kind_from_tipo(tipo_raw: str) -> str:
        """
        Clasifica variantes de 'Video'/'Reel' como reels y variantes de 'Imagen'/'Álbum' como posts.

        No cambia el valor original que se escribe en el Excel; solo se usa para decidir scraping/validación.
        """
        if not isinstance(tipo_raw, str):
            return "other"
        s = tipo_raw.strip()
        if not s:
            return "other"

        # Normaliza: quita acentos y pasa a minúsculas para aceptar cualquier forma de escritura.
        s_norm = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii").lower()

        # Reels / Video
        if "video" in s_norm or "reel" in s_norm:
            return "reel"

        # Posts: Imagen / Album
        if "imagen" in s_norm or "image" in s_norm:
            return "post"
        if "album" in s_norm:
            return "post"

        return "other"

    def build_output_row(item, link_construido, scraped_data, invalid_from_validation=False):
        id_publicacion = item['id_publicacion']
        usuario = item['usuario']
        cuenta = item['cuenta']
        tipo = item['tipo']
        texto = item['texto']
        fecha = item['fecha']
        scan_comments = item['scan_comments']
        scan_reactions = item['scan_reactions']
        scan_plays = item['scan_plays']

        publicacion_eliminada = "SI" if invalid_from_validation else "NO"

        real_comments = 0
        real_reactions = 0
        real_plays = 0

        if scraped_data:
            real_comments = scraped_data.get('comments', 0)
            real_reactions = scraped_data.get('likes', 0)
            # PostFinder usa 'views', ReelFinder usa 'plays'
            real_plays = scraped_data.get('views') if scraped_data.get('views') is not None else scraped_data.get('plays', 0)
        elif invalid_from_validation:
            # Criterio: inválidos por existencia -> métricas en 0
            real_comments = 0
            real_reactions = 0
            real_plays = 0

        if invalid_from_validation:
            # Criterio: inválido por existencia -> marcado inconsistente automáticamente
            es_inconsistente = True
        else:
            # Comparación y Umbral (5%)
            diff_c = calculate_diff(real_comments, scan_comments)
            diff_r = calculate_diff(real_reactions, scan_reactions)
            diff_p = calculate_diff(real_plays, scan_plays)
            es_inconsistente = (diff_c > 5) or (diff_r > 5) or (diff_p > 5)

        return {
            'index_original': item['index_original'],
            'Id Publicación': id_publicacion,
            'Usuario Cuenta': cuenta if cuenta and str(cuenta).lower() != 'nan' else usuario,
            'Link': link_construido,
            'Texto': texto,
            'Tipo publicación': tipo,
            'Fecha publicación': fecha,
            'SCAN: Comentarios': scan_comments,
            'SCAN: Reacciones': scan_reactions,
            'SCAN: Reproducciones': scan_plays,
            'REDES SOCIALES: Comentarios': real_comments,
            'REDES SOCIALES: Reacciones': real_reactions,
            'REDES SOCIALES: Reproducciones': real_plays,
            '__inconsistent__': es_inconsistente, # Columna oculta para formato
            'PUBLICACION ELIMINADA': publicacion_eliminada
        }

    # Procesamiento por usuario (agrupación)
    def save_partial_results(current_results):
        results_sorted = sorted(current_results, key=lambda x: x.get('index_original', 0))
        output_df = pd.DataFrame(results_sorted)

        if 'Id Publicación' in output_df.columns:
            output_df['Id Publicación'] = output_df['Id Publicación'].astype(str)

        with pd.ExcelWriter(output_final, engine='openpyxl') as writer:
            output_df.drop(columns=['__inconsistent__', 'index_original']).to_excel(
                writer, index=False, sheet_name='Comparativa'
            )

        # Aplicar formato condicional con openpyxl
        wb = load_workbook(output_final)
        ws = wb['Comparativa'] if 'Comparativa' in wb.sheetnames else wb.active
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        for i, res in enumerate(results_sorted):
            if res.get('__inconsistent__'):
                # i + 2 porque el Excel empieza en 1 y la fila 1 es el header
                ws.cell(row=i + 2, column=1).fill = red_fill

        wb.save(output_final)

    for usuario, user_df in df.groupby('Usuario', sort=False):
        usuario = str(usuario)
        logging.info(f"--- Procesando usuario {usuario} ({len(user_df)} filas) ---")

        # Preparar ítems para validación y scraping
        rows_to_process = []
        try:
            for _, row in user_df.iterrows():
                try:
                    id_publicacion = str(row.get('Id Publicación', '')).split('.')[0] 
                    cuenta = str(row.get('Cuenta', ''))
                    tipo = str(row.get('Tipo publicación', ''))
                    texto = str(row.get('Texto', ''))
                    fecha = str(row.get('Fecha publicación', ''))
                    enlace = row.get('Enlace', '')

                    # Métricas SCAN (del Excel original)
                    scan_comments = row.get('Comentarios', 0)
                    scan_reactions = row.get('Reacciones', 0)
                    scan_plays = row.get('Reproducciones', 0)

                    shortcode = extract_shortcode(enlace)
                    kind = infer_kind_from_tipo(tipo)

                    rows_to_process.append({
                        'index_original': int(row.get('index_original')),
                        'usuario': usuario,
                        'cuenta': cuenta,
                        'tipo': tipo,
                        'texto': texto,
                        'fecha': fecha,
                        'enlace': enlace,
                        'shortcode': shortcode,
                        'kind': kind,
                        'id_publicacion': id_publicacion,
                        'scan_comments': scan_comments,
                        'scan_reactions': scan_reactions,
                        'scan_plays': scan_plays,
                    })
                except Exception as e:
                    # Si una fila falla al preparar el item, evitamos romper todo el usuario
                    logging.error(f"Error preparando fila para usuario {usuario}: {e}")
        except Exception as e:
            logging.error(f"Error preparando datos para usuario {usuario}: {e}")

        # Shortcodes para validación batch (se deduplica para rendimiento)
        post_shortcodes = [it['shortcode'] for it in rows_to_process if it['kind'] == 'post' and it['shortcode']]
        reel_shortcodes = [it['shortcode'] for it in rows_to_process if it['kind'] == 'reel' and it['shortcode']]

        # Mantener orden de aparición al deduplicar
        unique_post_shortcodes = list(dict.fromkeys(post_shortcodes))
        unique_reel_shortcodes = list(dict.fromkeys(reel_shortcodes))

        post_exists_map = {}
        reel_exists_map = {}
        post_validation_ok = True
        reel_validation_ok = True

        logging.info(
            f"Validando {len(unique_post_shortcodes)} posts y {len(unique_reel_shortcodes)} reels para usuario {usuario}..."
        )

        # Validación batch por tipo (con un solo navegador)
        try:
            if unique_post_shortcodes:
                post_exists_map = validar_existencia_playwright_batch(
                    unique_post_shortcodes, tipo="post", auth_file="auth.json"
                )
        except Exception as e:
            logging.error(f"Error validando existencia de posts para usuario {usuario}: {e}")
            post_exists_map = {}
            post_validation_ok = False

        try:
            if unique_reel_shortcodes:
                reel_exists_map = validar_existencia_playwright_batch(
                    unique_reel_shortcodes, tipo="reel", auth_file="auth.json"
                )
        except Exception as e:
            logging.error(f"Error validando existencia de reels para usuario {usuario}: {e}")
            reel_exists_map = {}
            reel_validation_ok = False

        # Logs de existencia
        post_exist_count = sum(1 for v in post_exists_map.values() if v)
        post_nonexist_count = len(post_exists_map) - post_exist_count
        reel_exist_count = sum(1 for v in reel_exists_map.values() if v)
        reel_nonexist_count = len(reel_exists_map) - reel_exist_count

        logging.info(
            f"Existencia - Posts: {post_exist_count} existen / {post_nonexist_count} no existen; "
            f"Reels: {reel_exist_count} existen / {reel_nonexist_count} no existen"
        )

        # Caches para evitar scrapear múltiples veces el mismo shortcode
        post_scrape_cache = {}
        reel_scrape_cache = {}

        if post_validation_ok:
            post_valid_rows_count = sum(
                1 for it in rows_to_process
                if it['kind'] == 'post' and it['shortcode'] and post_exists_map.get(it['shortcode']) is True
            )
            post_valid_shortcodes_count = sum(1 for sc in unique_post_shortcodes if post_exists_map.get(sc) is True)
        else:
            post_valid_rows_count = sum(1 for it in rows_to_process if it['kind'] == 'post' and it['shortcode'])
            post_valid_shortcodes_count = len(unique_post_shortcodes)

        if reel_validation_ok:
            reel_valid_rows_count = sum(
                1 for it in rows_to_process
                if it['kind'] == 'reel' and it['shortcode'] and reel_exists_map.get(it['shortcode']) is True
            )
            reel_valid_shortcodes_count = sum(1 for sc in unique_reel_shortcodes if reel_exists_map.get(sc) is True)
        else:
            reel_valid_rows_count = sum(1 for it in rows_to_process if it['kind'] == 'reel' and it['shortcode'])
            reel_valid_shortcodes_count = len(unique_reel_shortcodes)

        processed_rows_total = 0

        # Lista de shortcodes a scrapear (únicos por tipo)
        post_shortcodes_to_scrape = (
            [sc for sc in unique_post_shortcodes if post_exists_map.get(sc) is True]
            if post_validation_ok
            else list(unique_post_shortcodes)
        )
        reel_shortcodes_to_scrape = (
            [sc for sc in unique_reel_shortcodes if reel_exists_map.get(sc) is True]
            if reel_validation_ok
            else list(unique_reel_shortcodes)
        )

        # Scraping optimizado: 1 navegador para posts y 1 para reels, reutilizando la misma page.
        post_profile_url = f"https://www.instagram.com/{usuario}/"
        reel_profile_url = f"https://www.instagram.com/{usuario}/reels/"

        if post_shortcodes_to_scrape:
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=False)
                    context = browser.new_context(storage_state="auth.json")
                    page = context.new_page()
                    try:
                        try:
                            scraped_map = post_finder.buscar_posts_multiples_en_pagina(
                                post_profile_url, post_shortcodes_to_scrape, page
                            )
                        except Exception as e:
                            logging.error(f"Error scraping posts batch ({usuario}): {e}")
                            scraped_map = {}

                        post_scrape_cache.update(scraped_map)
                    finally:
                        browser.close()
            except Exception as e:
                logging.error(f"Error abriendo Playwright para posts ({usuario}): {e}")

        if reel_shortcodes_to_scrape:
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=False)
                    context = browser.new_context(storage_state="auth.json")
                    page = context.new_page()
                    try:
                        try:
                            scraped_map = reel_finder.buscar_reels_multiples_en_pagina(
                                reel_profile_url, reel_shortcodes_to_scrape, page
                            )
                        except Exception as e:
                            logging.error(f"Error scraping reels batch ({usuario}): {e}")
                            scraped_map = {}

                        reel_scrape_cache.update(scraped_map)
                    finally:
                        browser.close()
            except Exception as e:
                logging.error(f"Error abriendo Playwright para reels ({usuario}): {e}")

        processed_scrape_calls = len(post_shortcodes_to_scrape) + len(reel_shortcodes_to_scrape)
        post_scraped_success = sum(1 for v in post_scrape_cache.values() if v)
        reel_scraped_success = sum(1 for v in reel_scrape_cache.values() if v)

        # Armar salida recorriendo filas y usando caches
        for item in rows_to_process:
            shortcode = item['shortcode']

            link_construido = ""
            if item['kind'] == 'reel':
                link_construido = reel_profile_url
            elif item['kind'] == 'post':
                link_construido = post_profile_url

            invalid_from_validation = False
            scraped_data = None

            # Determinar invalidación y datos desde cache
            if item['kind'] == 'post' and shortcode:
                if post_validation_ok:
                    exists = post_exists_map.get(shortcode)
                    if exists is False:
                        invalid_from_validation = True
                    elif exists is True:
                        scraped_data = post_scrape_cache.get(shortcode)
                else:
                    # Fallback: si validación falló, ya intentamos scrapear todo lo único.
                    scraped_data = post_scrape_cache.get(shortcode)
            elif item['kind'] == 'reel' and shortcode:
                if reel_validation_ok:
                    exists = reel_exists_map.get(shortcode)
                    if exists is False:
                        invalid_from_validation = True
                    elif exists is True:
                        scraped_data = reel_scrape_cache.get(shortcode)
                else:
                    scraped_data = reel_scrape_cache.get(shortcode)

            try:
                results.append(
                    build_output_row(
                        item,
                        link_construido,
                        scraped_data,
                        invalid_from_validation=invalid_from_validation
                    )
                )
                processed_rows_total += 1
            except Exception as e:
                logging.error(f"Error armando salida para usuario {usuario}: {e}")

        logging.info(
            f"Usuario {usuario}: filas válidas (por existencias) posts={post_valid_rows_count}, reels={reel_valid_rows_count}; "
            f"shortcodes a scrapear posts={len(post_shortcodes_to_scrape)}, reels={len(reel_shortcodes_to_scrape)}; "
            f"scrape exitoso posts={post_scraped_success}, reels={reel_scraped_success}; "
            f"llamadas de scraping ejecutadas={processed_scrape_calls}; filas procesadas={processed_rows_total}"
        )

        # Autoguardado: al terminar cada usuario, se reescribe el Excel manteniendo el orden.
        try:
            save_partial_results(results)
            logging.info(f"Autoguardado completado (usuario {usuario}). Filas actuales: {len(results)}")
        except Exception as e:
            logging.error(f"Error en autoguardado para usuario {usuario}: {e}")

    # Guardado final por seguridad (en caso de que el último autoguardado haya fallado).
    try:
        save_partial_results(results)
    except Exception as e:
        logging.error(f"Error guardando resultado final: {e}")
    logging.info(f"¡Proceso completado! Archivo generado: {output_final}")

if __name__ == "__main__":
    main()
